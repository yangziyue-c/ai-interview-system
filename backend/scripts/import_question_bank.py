"""题库导入脚本：把仓库根目录 题库/*.xlsx（v13）导入 questions 表

运行方式（必须在 backend 目录下）：
    cd backend
    <ai_interview 的 python.exe> -m scripts.import_question_bank

处理规则：
1. 「所属岗位」中文名 → position_code 的映射见 POSITION_MAP（本脚本是题库
   xlsx 格式的唯一消费者，映射在此维护；列名漂移时该岗位题进入 skipped 告警）；
2. 题干中的「【岗位软技能考察：X】」元信息剥离到 soft_skill_tag 列，
   保证题干可直接读给候选人；
3. 表头与列数强校验（与 xlsx 列序漂移即刻报错，不做静默错位）；
   大类/难度/阶段按受控词表校验，非法值跳过并汇总告警；
4. 幂等：按 (position_code, question_no) 更新或插入，可重复执行；
5. 题目编号跨岗位重复（每岗都有 tech_001），复合键唯一约束保证不冲突。
"""
import asyncio
import re
import sys
from pathlib import Path

# Windows 控制台默认 GBK 编码，print 中文会抛 UnicodeEncodeError；
# stdout 可能不是 TextIOWrapper（管道/嵌入环境），hasattr 守卫防止导入即崩
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from sqlalchemy import select

from app.database import async_session, init_db
from app.models import Question
from app.models.question import QUESTION_CATEGORIES, QUESTION_DIFFICULTIES, QUESTION_STAGES

# 题库源目录：仓库根目录/题库
BANK_DIR = Path(__file__).resolve().parents[2] / "题库"

# 题库 xlsx「所属岗位」列中文名 → 岗位 code
# ⚠️ 与 xlsx 逐字一致（「Java后端」无空格）；漂移时该岗位题进入 skipped 告警
POSITION_MAP = {
    "Java后端": "backend",
    "Web前端": "frontend",
    "软件测试开发": "test_engineer",
}

# 题库 15 列（列顺序固定，与 xlsx 表头逐字一致，导入前强校验）
HEADERS = [
    "题目编号", "所属岗位", "大类", "题目分类", "难度等级", "面试问题", "得分点",
    "追问触发条件", "参考答案", "备注", "面试阶段", "建议用时(分钟)", "阶段顺序",
    "替代回答方向", "优秀回答范例",
]

# 受控词表（与 app/models/question.py 的常量一致）
_VOCABULARY = (
    ("大类", QUESTION_CATEGORIES),
    ("难度等级", QUESTION_DIFFICULTIES),
    ("面试阶段", QUESTION_STAGES),
)

_TAG_RE = re.compile(r"【岗位软技能考察：([^】]*)】")


def strip_soft_skill_tag(text: str) -> tuple[str, str]:
    """把题干里的软技能考察标签剥离出来，返回 (干净题干, 标签文本)"""
    tag = ""
    m = _TAG_RE.search(text)
    if m:
        tag = m.group(1).strip()
    cleaned = _TAG_RE.sub("", text)
    # 标签通常独占一段，剥离后清理多余空行
    cleaned = re.sub(r"\n{2,}", "\n", cleaned).strip()
    return cleaned, tag


def _text(value) -> str:
    """单元格转字符串，None/空 → ''"""
    if value is None:
        return ""
    return str(value).strip()


def _int(value, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


async def import_bank() -> dict:
    """导入全部题库文件，返回统计信息"""
    xlsx_files = sorted(BANK_DIR.glob("*.xlsx"))
    if not xlsx_files:
        raise FileNotFoundError(f"题库目录未找到 xlsx 文件：{BANK_DIR}")

    await init_db()

    stats = {"files": len(xlsx_files), "added": 0, "updated": 0, "tags_stripped": 0,
             "per_position": {}, "skipped": {}}

    async with async_session() as session:
        # 先加载库内全部题目，按复合键索引，避免逐条查询
        existing = {
            (q.position_code, q.question_no): q
            for q in (await session.scalars(select(Question))).all()
        }

        for xlsx_file in xlsx_files:
            wb = load_workbook(xlsx_file, data_only=True, read_only=True)
            if "面试题库" not in wb.sheetnames:
                print(f"⚠️ 跳过 {xlsx_file.name}：缺少「面试题库」sheet")
                wb.close()
                continue
            ws = wb["面试题库"]

            # 表头强校验：列名/列序与 HEADERS 不一致立即报错，防止字段静默错位入库
            actual_headers = [_text(c.value) for c in next(ws.iter_rows(max_row=1))]
            if actual_headers != HEADERS:
                wb.close()
                raise ValueError(f"{xlsx_file.name} 表头与预期不符：\n  实际: {actual_headers}\n  预期: {HEADERS}")

            batch_keys: set[tuple[str, str]] = set()  # 本批次已处理的复合键（检测重复编号）
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) != len(HEADERS):
                    stats["skipped"]["列数不符"] = stats["skipped"].get("列数不符", 0) + 1
                    continue
                data = dict(zip(HEADERS, row))

                question_no = _text(data.get("题目编号"))
                if not question_no:
                    stats["skipped"]["缺题目编号"] = stats["skipped"].get("缺题目编号", 0) + 1
                    continue

                # 受控词表校验：非法值跳过并汇总，防止脏数据入库
                invalid = [f"{label}={_text(data.get(label))}" for label, vocab in _VOCABULARY
                           if _text(data.get(label)) not in vocab]
                if invalid:
                    for item in invalid:
                        stats["skipped"][item] = stats["skipped"].get(item, 0) + 1
                    continue

                position_code = POSITION_MAP.get(_text(data.get("所属岗位")))
                if position_code is None:
                    label = f"岗位={_text(data.get('所属岗位'))}"
                    stats["skipped"][label] = stats["skipped"].get(label, 0) + 1
                    continue

                question, tag = strip_soft_skill_tag(_text(data.get("面试问题")))
                if tag:
                    stats["tags_stripped"] += 1

                values = dict(
                    position_code=position_code,
                    question_no=question_no,
                    category=_text(data.get("大类")),
                    sub_category=_text(data.get("题目分类")),
                    difficulty=_text(data.get("难度等级")),
                    question=question,
                    soft_skill_tag=tag,
                    score_points=_text(data.get("得分点")),
                    follow_up_triggers=_text(data.get("追问触发条件")),
                    reference_answer=_text(data.get("参考答案")),
                    note=_text(data.get("备注")),
                    interview_stage=_text(data.get("面试阶段")),
                    stage_order=_int(data.get("阶段顺序")),
                    suggested_minutes=_int(data.get("建议用时(分钟)")),
                    alternative_directions=_text(data.get("替代回答方向")),
                    excellent_example=_text(data.get("优秀回答范例")),
                )

                key = (position_code, question_no)
                if key in existing:
                    if key in batch_keys:
                        print(f"⚠️ 重复题号（本批次覆盖）：{xlsx_file.name} {question_no}")
                    for field, value in values.items():
                        setattr(existing[key], field, value)
                    stats["updated"] += 1
                else:
                    question_obj = Question(**values)
                    session.add(question_obj)
                    existing[key] = question_obj  # 登记，防同批次重复编号二次 add 触发唯一约束
                    stats["added"] += 1
                batch_keys.add(key)

                stats["per_position"][position_code] = stats["per_position"].get(position_code, 0) + 1

            wb.close()

        await session.commit()

    return stats


def main() -> None:
    print("=" * 60)
    print("题库导入：题库/*.xlsx → questions 表")
    print(f"源目录：{BANK_DIR}")
    print("=" * 60)
    stats = asyncio.run(import_bank())
    print(f"✅ 导入完成：{stats['files']} 个文件，新增 {stats['added']} 题，更新 {stats['updated']} 题")
    print(f"   剥离软技能标签：{stats['tags_stripped']} 题")
    for code, count in stats["per_position"].items():
        print(f"   {code}: {count} 题")
    if stats["skipped"]:
        print(f"⚠️ 校验失败已跳过：{stats['skipped']}")


if __name__ == "__main__":
    main()
